"""استيراد بيانات الطرق من ملف Excel (إضافة فقط — لا يعدّل/يحذف أي بيانات موجودة).

تشغيل:
  python manage.py import_excel_roads --excel "...\\العيوب3.xlsx" --photos "...\\Photo"            # معاينة (dry-run)
  python manage.py import_excel_roads --excel "..." --photos "..." --commit                          # تنفيذ فعلي
"""
import os
import re
from django.core.management.base import BaseCommand, CommandError
from django.core.files.base import ContentFile
from django.db import transaction
from openpyxl import load_workbook
from bridges.models import Road, RoadDefect, RoadDefectImage

DIRMAP = {"اتجاه جدة": "jeddah", "اتجاه مكة": "mecca"}
SEV = {"منخفض": "low", "متوسط": "medium", "حرج": "critical"}
RANK = {"low": 1, "medium": 2, "critical": 3}
RANK2 = {1: "low", 2: "medium", 3: "critical"}
CLS = {"ملاحظات اسفلت": "اسفلت", "ملاحظات أصول جانبي الطريق": "أصول جانبي الطريق"}
COL = {"photo": "رقم الصورة", "type": "نوع العيب", "cls": "تصنيف الملاحظة",
       "seg": "رقم المقطع", "treat": "حالة المعالجة", "lat": "خط_ العرض",
       "lng": "خط_الطول", "sev": "درجة الخطورة"}


def norm_segment(s):
    s = (str(s) if s is not None else "").strip()
    if s == "" or re.fullmatch(r"_+", s):
        return "بدون رقم"
    return s


def photo_index(folder):
    idx = {}
    for f in os.listdir(folder):
        idx[f.lower()] = f
    return idx


def find_photo(token, idx):
    t = token.strip().strip("ـ").strip()  # إزالة رمز التطويل والمسافات
    if not t:
        return None
    for ext in ("png", "PNG", "jpg", "jpeg", "JPG", "JPEG"):
        k = f"{t}.{ext}".lower()
        if k in idx:
            return idx[k]
    m = re.match(r"[Mm][-_ ]?0*(\d+)$", t)
    if m:
        n = int(m.group(1))
        for cand in (f"m_{n:03d}.png", f"m_{n}.png", f"M-{n:03d}.PNG", f"M-{n}.PNG"):
            if cand.lower() in idx:
                return idx[cand.lower()]
    return None


class Command(BaseCommand):
    help = "استيراد بيانات الطرق من Excel (إضافة فقط)"

    def add_arguments(self, p):
        p.add_argument("--excel", required=True)
        p.add_argument("--photos", required=True)
        p.add_argument("--commit", action="store_true", help="تنفيذ الكتابة فعليًا (بدونها معاينة فقط)")

    def handle(self, *args, **o):
        excel, photos, commit = o["excel"], o["photos"], o["commit"]
        if not os.path.exists(excel):
            raise CommandError(f"الملف غير موجود: {excel}")
        if not os.path.isdir(photos):
            raise CommandError(f"فولدر الصور غير موجود: {photos}")
        idx = photo_index(photos)

        wb = load_workbook(excel, read_only=True, data_only=True)
        rows = []
        for sn in wb.sheetnames:
            data = list(wb[sn].iter_rows(values_only=True))
            if not data:
                continue
            hdr = data[0]
            ci = {key: hdr.index(col) for key, col in COL.items() if col in hdr}
            for r in data[1:]:
                if r is None or all(c is None for c in r):
                    continue
                rows.append({
                    "dir": DIRMAP.get(sn, sn),
                    "seg": norm_segment(r[ci["seg"]]),
                    "type": (r[ci["type"]] or "").strip(),
                    "cls": CLS.get((r[ci["cls"]] or "").strip(), (r[ci["cls"]] or "").strip()),
                    "sev": SEV.get((r[ci["sev"]] or "").strip(), "low"),
                    "lat": r[ci["lat"]], "lng": r[ci["lng"]],
                    "photos": [t.strip() for t in str(r[ci["photo"]] or "").split("&") if t.strip()],
                })

        # تجميع حسب (اتجاه، مقطع)
        groups = {}
        for row in rows:
            groups.setdefault((row["dir"], row["seg"]), []).append(row)

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\n=== خطة الاستيراد (إجمالي {len(rows)} حالة في {len(groups)} مقطع) ==="))
        img_found = img_missing = 0
        missing_tokens = []
        for (dirn, seg), grp in sorted(groups.items()):
            existing = Road.objects.filter(segment=seg, direction=dirn).first()
            worst = RANK2[max(RANK[g["sev"]] for g in grp)]
            tag = f"موجود (id={existing.id}, status={existing.status} — لن يتغيّر)" if existing else f"جديد (status={worst})"
            for g in grp:
                for tok in g["photos"]:
                    if find_photo(tok, idx):
                        img_found += 1
                    else:
                        img_missing += 1
                        missing_tokens.append(tok)
            self.stdout.write(f"  [{dirn}] {seg}: {len(grp)} حالة → {tag}")
        self.stdout.write(f"\nالصور: {img_found} موجودة، {img_missing} مفقودة {missing_tokens[:10] if missing_tokens else ''}")

        if not commit:
            self.stdout.write(self.style.WARNING("\n[معاينة فقط] لم تتم أي كتابة. أضِف --commit للتنفيذ."))
            return

        # ===== التنفيذ الفعلي (إضافة فقط) =====
        created_roads = created_defects = created_imgs = 0
        with transaction.atomic():
            for (dirn, seg), grp in groups.items():
                worst = RANK2[max(RANK[g["sev"]] for g in grp)]
                road, was_created = Road.objects.get_or_create(
                    segment=seg, direction=dirn, defaults={"status": worst})
                if was_created:
                    created_roads += 1
                for g in grp:
                    d = RoadDefect.objects.create(
                        road=road, title=g["type"], status=g["sev"],
                        observation=g["cls"], treatment_type="untreated",
                        description="", lat=float(g["lat"]), lng=float(g["lng"]),
                    )
                    created_defects += 1
                    for tok in g["photos"]:
                        fn = find_photo(tok, idx)
                        if not fn:
                            continue
                        with open(os.path.join(photos, fn), "rb") as fh:
                            di = RoadDefectImage(defect=d, caption=tok)
                            di.image.save(fn, ContentFile(fh.read()), save=True)
                            created_imgs += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nتم: {created_roads} مقطع جديد · {created_defects} حالة · {created_imgs} صورة. (لم يُمسّ أي بيانات موجودة)"))
